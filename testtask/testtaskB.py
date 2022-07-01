import os
import openpyxl
wb = openpyxl.load_workbook(filename=r'C:\Users\HP\Downloads\testtask (1)\testtask\result.xlsx')
sheet = wb['result']

print(os.path.abspath(__file__))
current_address = os.path.dirname(os.path.abspath(__file__))

filelist = []

for root, dirs, files in os.walk(current_address):
    for file in files:
        filelist.append(os.path.join(root, file))
i = 1
for name in filelist:
    sheet.cell(row=i, column=1).value = os.path.split(os.path.split(name)[0])[1]
    sheet.cell(row=i, column=2).value = os.path.splitext(os.path.basename(name))[0]
    sheet.cell(row=i, column=3).value = str(os.path.splitext(os.path.basename(name))[1])[1:]
    i+=1
    print(os.path.split(os.path.split(name)[0])[1])
    print(os.path.splitext(os.path.basename(name))[0])
    print(str(os.path.splitext(os.path.basename(name))[1])[1:])

wb.save(r'C:\Users\HP\Downloads\testtask (1)\testtask\result.xlsx')
    # print(os.path.split(os.path.split(name)[0])[1])
    # print(os.path.splitext(os.path.basename(name))[0])
    # print(str(os.path.splitext(os.path.basename(name))[1])[1:])